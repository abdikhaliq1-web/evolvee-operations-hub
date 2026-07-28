from io import BytesIO

from django.db.models import Count, Q, QuerySet
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from partners.models import Partner, ProgramActivity, SaleStatus
from partners.utils.commission import get_partner_stats


PARTNER_HEADERS = [
    "Creator Name",
    "Creator ID",
    "Discount Code",
    "Email",
    "Social Handle",
    "Status",
    "City",
    "State / Region",
    "Country",
    "Continent",
    "Payment Method",
    "Payment Details",
    "Payout Ready",
    "QR Scans",
    "Conversions",
    "Conversion Rate %",
    "Revenue ($)",
    "Commission ($)",
    "Pending Commission ($)",
    "Trackable Link",
    "Store Referral Link",
    "QR Code",
]

ACTIVITY_HEADERS = [
    "Date",
    "Event Type",
    "Partner",
    "Partner Code",
    "Description",
    "User Email",
    "IP Address",
]


def _auto_column_widths(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_partners_workbook(queryset: QuerySet | None = None) -> BytesIO:
    partners = queryset if queryset is not None else Partner.objects.all()
    partners = (
        partners.select_related("user")
        .annotate(
            click_count=Count("clicks"),
            conversion_count=Count("clicks", filter=Q(clicks__converted=True)),
            sale_count=Count("sales", filter=Q(sales__status=SaleStatus.APPROVED)),
        )
        .order_by("partner_name")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Creators"

    header_font = Font(bold=True)
    for col, header in enumerate(PARTNER_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    row_idx = 2
    qr_col = PARTNER_HEADERS.index("QR Code") + 1
    for partner in partners:
        stats = get_partner_stats(partner)
        ws.row_dimensions[row_idx].height = 95

        values = [
            partner.partner_name,
            partner.partner_code or "—",
            partner.discount_code or "—",
            partner.user.email,
            partner.social_handle or "—",
            partner.get_status_display(),
            partner.city or "—",
            partner.region or "—",
            partner.country or "—",
            partner.continent or "—",
            partner.get_payment_method_display() if partner.payment_method else "—",
            partner.payment_details or "—",
            "Yes" if partner.has_payment_details else "No",
            stats["total_clicks"],
            stats["total_conversions"],
            stats["conversion_rate"],
            float(partner.total_sales),
            float(partner.total_commission_earned),
            float(stats["pending_commission"]),
            partner.tracking_url or "—",
            partner.referral_url or "—",
            "",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        if partner.qr_code_image:
            try:
                img = XLImage(partner.qr_code_image.path)
                img.width = 80
                img.height = 80
                ws.add_image(img, f"{get_column_letter(qr_col)}{row_idx}")
            except (FileNotFoundError, OSError):
                ws.cell(row=row_idx, column=qr_col, value="QR file missing")

        row_idx += 1

    _auto_column_widths(
        ws,
        {
            1: 22,
            2: 14,
            3: 14,
            4: 28,
            5: 18,
            6: 12,
            7: 16,
            8: 18,
            9: 12,
            10: 16,
            11: 16,
            12: 32,
            13: 12,
            14: 10,
            15: 12,
            16: 14,
            17: 12,
            18: 14,
            19: 18,
            20: 36,
            21: 36,
            22: 14,
        },
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_activity_workbook(queryset: QuerySet | None = None) -> BytesIO:
    activities = queryset if queryset is not None else ProgramActivity.objects.all()
    activities = activities.select_related("partner", "user").order_by("-created_at")

    wb = Workbook()
    ws = wb.active
    ws.title = "Program Activity"

    header_font = Font(bold=True)
    for col, header in enumerate(ACTIVITY_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    for row_idx, activity in enumerate(activities, start=2):
        values = [
            activity.created_at.strftime("%Y-%m-%d %H:%M"),
            activity.get_event_type_display(),
            activity.partner.partner_name if activity.partner else "—",
            activity.partner.partner_code if activity.partner else "—",
            activity.description or "—",
            activity.user.email if activity.user else "—",
            activity.ip_address or "—",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=value)

    _auto_column_widths(ws, {1: 18, 2: 22, 3: 22, 4: 14, 5: 40, 6: 28, 7: 16})

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
